"""
Turns one parsed session (the dict returned by parser.parse_rinex_file)
into the same 3-sheet Excel workbook style used earlier in this project
(Metadata, Satellite Summary, Epoch Time Series).

This is deliberately fed by parse_rinex_file()'s output directly, so the
SAME parsing logic that populates the database also produces the Excel
export -- there is only one source of truth for "what the data is",
whether the destination is a database row or a spreadsheet.
"""
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
good_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
part_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
poor_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")


def _ecef_to_geodetic(x, y, z):
    a = 6378137.0
    f = 1 / 298.257223563
    e2 = f * (2 - f)
    lon = math.atan2(y, x)
    p = math.sqrt(x*x + y*y)
    lat = math.atan2(z, p * (1 - e2))
    for _ in range(10):
        N = a / math.sqrt(1 - e2 * math.sin(lat)**2)
        h = p / math.cos(lat) - N
        lat = math.atan2(z, p * (1 - e2 * N / (N + h)))
    N = a / math.sqrt(1 - e2 * math.sin(lat)**2)
    h = p / math.cos(lat) - N
    return math.degrees(lat), math.degrees(lon), h


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _quality_fill(q):
    if q in ("Excellent", "Good"):
        return good_fill
    if q.startswith("Partial"):
        return part_fill
    return poor_fill


def export_session_to_excel(result: dict, source_filename: str, out_path: str):
    """
    result: the dict returned by app.parser.parse_rinex_file()
    source_filename: original .rnx filename, for the Metadata sheet
    out_path: where to write the .xlsx file
    """
    meta = result["meta"]
    lat, lon, h = (None, None, None)
    if meta.get("x") is not None:
        lat, lon, h = _ecef_to_geodetic(meta["x"], meta["y"], meta["z"])

    wb = Workbook()

    # ---- Sheet 1: Metadata ----
    ws1 = wb.active
    ws1.title = "Metadata"
    ws1["A1"] = f"GNSS Session Metadata — {meta.get('marker') or 'Unknown station'}"
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells("A1:B1")
    ws1.append([])
    ws1.append(["Field", "Value"])
    _style_header(ws1, 3, 2)
    rows = [
        ("Source File", source_filename),
        ("Station Marker Name", meta.get("marker")),
        ("Receiver Type", meta.get("receiver")),
        ("Antenna Type", meta.get("antenna")),
        ("Approx Position X (m)", meta.get("x")),
        ("Approx Position Y (m)", meta.get("y")),
        ("Approx Position Z (m)", meta.get("z")),
        ("Latitude (deg, WGS84)", round(lat, 6) if lat else None),
        ("Longitude (deg, WGS84)", round(lon, 6) if lon else None),
        ("Height (m)", round(h, 3) if h else None),
        ("Session Start (UTC)", meta.get("start").strftime("%Y-%m-%d %H:%M:%S") if meta.get("start") else None),
        ("Expected Epochs", result.get("expected_epochs")),
        ("Actual Epochs Parsed", len(result["epochs"])),
        ("Satellites Tracked", len(result["satellites"])),
    ]
    for r in rows:
        ws1.append(list(r))
    _autosize(ws1, [26, 34])
    ws1.freeze_panes = "A4"

    # ---- Sheet 2: Satellite Summary ----
    ws2 = wb.create_sheet("Satellite Summary")
    ws2["A1"] = "Satellite Tracking & Data Quality Summary"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:J1")
    ws2.append([])
    headers2 = ["PRN", "System", "Valid Epochs", "Completeness %", "Avg SNR (dB-Hz)",
                "Cycle Slip Flags", "Quality", "Multipath MP1 (m)", "Multipath MP2 (m)", "Iono Delay Est. L1 (m)"]
    ws2.append(headers2)
    _style_header(ws2, 3, len(headers2))
    for s in sorted(result["satellites"], key=lambda r: r["prn"]):
        ws2.append([
            s["prn"], s["system"], s["valid_epochs"], s["completeness_pct"], s["avg_snr"],
            s["cycle_slips"], s["quality"], s.get("multipath_mp1"), s.get("multipath_mp2"), s.get("iono_delay_l1"),
        ])
    for i, s in enumerate(sorted(result["satellites"], key=lambda r: r["prn"]), start=4):
        ws2.cell(row=i, column=7).fill = _quality_fill(s["quality"])
    _autosize(ws2, [10, 12, 14, 16, 16, 16, 26, 18, 18, 20])
    ws2.freeze_panes = "A4"

    # ---- Sheet 3: Epoch Time Series ----
    ws3 = wb.create_sheet("Epoch Time Series")
    ws3["A1"] = "Per-Epoch Satellite Counts & Average SNR (1 Hz)"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:L1")
    ws3.append([])
    headers3 = ["Epoch (UTC)", "Total Satellites",
                "GPS Count", "GPS Avg SNR", "GLONASS Count", "GLONASS Avg SNR",
                "Galileo Count", "Galileo Avg SNR", "BeiDou Count", "BeiDou Avg SNR",
                "SBAS Count", "SBAS Avg SNR"]
    ws3.append(headers3)
    _style_header(ws3, 3, len(headers3))
    for e in result["epochs"]:
        ws3.append([
            e["time"].strftime("%Y-%m-%d %H:%M:%S"), e["total"],
            e["G_count"], e["G_avg_snr"], e["R_count"], e["R_avg_snr"],
            e["E_count"], e["E_avg_snr"], e["C_count"], e["C_avg_snr"],
            e["S_count"], e["S_avg_snr"],
        ])
    _autosize(ws3, [20] + [14] * 11)
    ws3.freeze_panes = "A4"

    wb.save(out_path)
    return out_path
